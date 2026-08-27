#!/usr/bin/env python3
"""Gabungkan PDF bernama sama dari subfolder berbeda menjadi satu PDF.

Contoh struktur yang didukung:

    input/
      eklaim/0099R0010726V000032.pdf     <- 1
      rm/0099R0010726V000032.pdf         <- 2 (nama sama)
    result/0099R0010726V000032.pdf       <- gabungan 1+2

Cara pakai (dari folder proyek ini):

    python combine.py                      # input/ -> result/ (hanya duplikat)
    python combine.py --dry-run            # lihat rencana, tanpa menulis
    python combine.py --include-unique     # proses juga nama yang hanya 1 folder
    python combine.py --output "D:\\gabung"
    python combine.py --engine gs          # paksa Ghostscript

Engine gabung (--engine / env EKLAIM_COMBINE_ENGINE):
    auto  (default) : Ghostscript bila tersedia, fallback ke pypdf
    gs              : paksa Ghostscript (engine utama)
    pypdf           : paksa pypdf (opsional, fallback saja)

Ghostscript adalah engine utama (hasil lebih kecil, paling andal); pypdf
hanya dipakai sebagai fallback + validasi jumlah halaman. Install pypdf
bersifat opsional selama Ghostscript tersedia.

Sumber tidak pernah diubah — hanya dibaca.
"""
from __future__ import annotations

import argparse
import glob
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

PROJECT_DIR = Path(__file__).resolve().parent

try:
    from pypdf import PdfReader, PdfWriter
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

# ---------------------------------------------------------------- env
COMBINE_ENGINE = os.getenv("EKLAIM_COMBINE_ENGINE", "auto").strip().lower()
GS_EXECUTABLE = os.getenv("EKLAIM_GS_EXECUTABLE", "")

# ---------------------------------------------------------------- logging
def _setup_logging(log_dir: Path = PROJECT_DIR) -> logging.Logger:
    logger = logging.getLogger("combine")
    logger.setLevel(logging.DEBUG)
    for h in list(logger.handlers):
        logger.removeHandler(h)
    fmt = logging.Formatter("%(asctime)s %(levelname)-7s %(message)s", "%H:%M:%S")
    log_dir.mkdir(parents=True, exist_ok=True)
    fh = logging.FileHandler(log_dir / "combine.log", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


log = _setup_logging()

# ---------------------------------------------------------------- helpers
def _nat_key(s: str):
    """Natural sort: 'Tanggal 5' < 'Tanggal 25'."""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def find_pdfs(root: Path) -> list[Path]:
    return sorted(p for p in root.rglob("*")
                  if p.is_file() and p.suffix.lower() == ".pdf")


@dataclass(frozen=True)
class InputSources:
    """Sumber PDF dari folder induk lama atau dua folder langsung GUI."""

    folders: tuple[tuple[str, Path], ...]

    def files(self) -> list[Path]:
        return sorted(pdf for _, folder in self.folders for pdf in find_pdfs(folder))

    def source_name(self, path: Path) -> str:
        for name, folder in self.folders:
            try:
                path.relative_to(folder)
                return name
            except ValueError:
                continue
        return "(tidak diketahui)"

    def display_path(self, path: Path) -> str:
        for name, folder in self.folders:
            try:
                return str(Path(name) / path.relative_to(folder))
            except ValueError:
                continue
        return str(path)


def sources_from_root(root: Path) -> InputSources:
    """Mode CLI lama: setiap anak folder langsung menjadi sumber."""
    folders = tuple((folder.name, folder) for folder in sorted(root.iterdir()) if folder.is_dir())
    return InputSources(folders or ((root.name or "(root)", root),))


def group_by_name(files: list[Path]) -> dict[str, list[Path]]:
    groups: dict[str, list[Path]] = {}
    for f in files:
        groups.setdefault(f.name.lower(), []).append(f)
    return groups


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def page_count(path: Path) -> int:
    """Jumlah halaman: pypdf (bila ada), fallback hitung via Ghostscript."""
    if HAS_PYPDF:
        with open(path, "rb") as fh:
            return len(PdfReader(fh).pages)
    gs = find_gs()
    if not gs:
        return 0
    try:
        # -dNOSAFER: izinkan baca file lokal untuk penghitungan halaman
        cmd = [gs, "-q", "-dNOSAFER", "-dNODISPLAY", "-c",
               f"({path}) (r) file runpdfbegin pdfpagecount = quit"]
        out = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        return int(out.stdout.strip().splitlines()[-1])
    except Exception:
        return 0


# ---------------------------------------------------------------- engines
def find_gs() -> str | None:
    """Deteksi Ghostscript: env -> PATH -> folder install umum Windows.

    Utamakan executable asli (gswin64c/gs); tolak shim .cmd/.bat (gs.cmd)
    yang tidak bisa dipanggil langsung via subprocess.
    """
    if GS_EXECUTABLE and Path(GS_EXECUTABLE).exists():
        return GS_EXECUTABLE
    for name in ("gswin64c", "gswin32c", "gs"):
        found = shutil.which(name)
        if found and not found.lower().endswith((".cmd", ".bat")):
            return found
    for pattern in (
        r"C:\Program Files\gs\gs*\bin\gswin64c.exe",
        r"C:\Program Files (x86)\gs\gs*\bin\gswin64c.exe",
    ):
        hits = sorted(glob.glob(pattern), key=_nat_key, reverse=True)
        if hits:
            return hits[0]
    return None


def merge_pypdf(paths: list[Path], out_path: Path) -> int:
    writer = PdfWriter()
    for p in paths:
        with open(p, "rb") as fh:
            reader = PdfReader(fh)
            for page in reader.pages:
                writer.add_page(page)
    with open(out_path, "wb") as fh:
        writer.write(fh)
    return page_count(out_path)


def merge_gs(paths: list[Path], out_path: Path) -> int:
    gs = find_gs()
    if not gs:
        raise RuntimeError(
            "Ghostscript tidak ditemukan — set EKLAIM_GS_EXECUTABLE atau "
            "pasang di PATH (contoh: D:\\Program Files\\gs\\gs10.06.0\\bin\\gswin64c.exe)"
        )
    cmd = [
        gs, "-q", "-dNOPAUSE", "-dBATCH", "-sDEVICE=pdfwrite",
        "-dPDFSETTINGS=/prepress", f"-sOutputFile={out_path}",
    ] + [str(p) for p in paths]
    subprocess.run(cmd, check=True, capture_output=True)
    if not out_path.exists() or out_path.stat().st_size == 0:
        raise RuntimeError("Ghostscript tidak menghasilkan file output")
    return page_count(out_path)


def merge_any(paths: list[Path], out_path: Path, engine: str) -> int:
    """Gabung sesuai engine. auto = Ghostscript dulu, fallback pypdf."""
    if engine == "pypdf":
        if not HAS_PYPDF:
            raise RuntimeError(
                "pypdf tidak terpasang — install via: pip install pypdf"
            )
        return merge_pypdf(paths, out_path)

    if engine == "gs":
        return merge_gs(paths, out_path)  # error jelas bila gs tidak ada

    # auto: GS utama, pypdf fallback
    if find_gs():
        try:
            return merge_gs(paths, out_path)
        except Exception as exc:
            if HAS_PYPDF:
                log.warning("Ghostscript gagal (%s) — fallback ke pypdf.", exc)
                return merge_pypdf(paths, out_path)
            raise
    if HAS_PYPDF:
        log.warning("Ghostscript tidak terdeteksi — memakai pypdf.")
        return merge_pypdf(paths, out_path)
    raise RuntimeError(
        "Tidak ada engine tersedia: Ghostscript tidak ditemukan dan pypdf "
        "belum terpasang. Install salah satunya."
    )


# ---------------------------------------------------------------- verifikasi nama file vs NOSEP isi
SEP_RE = re.compile(r"\d{4}[A-Za-z]\d{7}[A-Za-z]\d{6}")
CHECK_CACHE_FILE = PROJECT_DIR / "combine_check_cache.json"


def _load_check_cache() -> dict:
    try:
        with open(CHECK_CACHE_FILE, encoding="utf-8") as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_check_cache(cache: dict) -> None:
    with open(CHECK_CACHE_FILE, "w", encoding="utf-8") as fh:
        json.dump(cache, fh, indent=1)


def _verify_one(p: Path, sources: InputSources) -> dict:
    """Cek nama file vs SEP yang ada di isi PDF.

    Status: ok | mismatch | no_sep (teks ada tapi tanpa pola SEP) |
    no_text (scan/tidak bisa dibaca).
    """
    rel = sources.display_path(p)
    name_sep = p.stem
    try:
        text = "".join((pg.extract_text() or "") for pg in PdfReader(str(p)).pages)
    except Exception:
        return {"file": rel, "path": str(p.resolve()), "nama_sep": name_sep, "isi_sep": "",
                "status": "no_text", "keterangan": "tidak bisa dibaca"}
    if not text.strip():
        return {"file": rel, "path": str(p.resolve()), "nama_sep": name_sep, "isi_sep": "",
                "status": "no_text", "keterangan": "PDF scan tanpa lapisan teks"}
    found = {m.upper() for m in SEP_RE.findall(text)}
    if not found:
        return {"file": rel, "path": str(p.resolve()), "nama_sep": name_sep, "isi_sep": "",
                "status": "no_sep", "keterangan": "tidak ada pola SEP di isi"}
    if name_sep.upper() in found:
        return {"file": rel, "path": str(p.resolve()), "nama_sep": name_sep, "isi_sep": name_sep.upper(),
                "status": "ok", "keterangan": ""}
    return {"file": rel, "path": str(p.resolve()), "nama_sep": name_sep, "isi_sep": ", ".join(sorted(found)[:6]),
            "status": "mismatch", "keterangan": "nama file != SEP di isi"}


def verify_names(sources: InputSources, force_verify: bool,
                 progress_callback: Callable[[str, int, int], None] | None = None) -> tuple[list[dict], dict]:
    """Verifikasi SEMUA PDF sumber; pakai cache (path,size,mtime) bila ada."""
    files = sources.files()
    cache = {} if force_verify else _load_check_cache()
    rows: list[dict] = []
    counts = {"ok": 0, "mismatch": 0, "no_sep": 0, "no_text": 0, "cached": 0}
    for index, p in enumerate(files, start=1):
        if progress_callback:
            progress_callback(f"Memeriksa nama dan SEP {index}/{len(files)}: {p.name}", index - 1, len(files))
        st = p.stat()
        key = str(p.resolve())  # normalisasi path: relatif/absolut konsisten di cache
        entry = cache.get(key)
        if (not force_verify and entry
                and entry.get("size") == st.st_size
                and entry.get("mtime") == st.st_mtime):
            row = entry["row"]
            counts["cached"] += 1
        else:
            row = _verify_one(p, sources)
            cache[key] = {"size": st.st_size, "mtime": st.st_mtime, "row": row}
        row.setdefault("path", str(p.resolve()))
        rows.append(row)
        if row["status"] in counts:
            counts[row["status"]] += 1
    _save_check_cache(cache)
    counts["total"] = len(files)
    if progress_callback:
        progress_callback("Pemeriksaan SEP selesai", len(files), len(files))
    return rows, counts


def _write_check_csv(rows: list[dict], report_dir: Path) -> None:
    import csv
    out = report_dir / "combine_check.csv"
    fieldnames = ["file", "nama_sep", "isi_sep", "status", "keterangan"]
    labels = {"file": "File", "nama_sep": "Nama File (SEP)",
              "isi_sep": "SEP di Isi", "status": "Status", "keterangan": "Keterangan"}
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writerow({k: labels[k] for k in fieldnames})
        writer.writerows(rows)


# ---------------------------------------------------------------- hasil
_STATUS_LABEL = {
    "merged": "Berhasil digabung",
    "copied": "Disalin (1 folder)",
    "identical": "File identik — disalin sekali",
    "skipped": "Nama unik — dilewati",
    "exists_skipped": "Output sudah ada — dilewati",
    "failed": "Gagal",
}
_RESULT_HEADERS = ["name", "nosep", "status", "keterangan", "pages", "detail"]
_RESULT_HEADER_LABEL = {
    "name": "Nama File", "nosep": "NOSEP", "status": "Status",
    "keterangan": "Keterangan", "pages": "Halaman", "detail": "Detail",
}


def _write_result(rows: list[dict], report_dir: Path) -> None:
    import csv
    out = report_dir / "combine_result.csv"
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_RESULT_HEADERS)
        writer.writerow({k: _RESULT_HEADER_LABEL[k] for k in _RESULT_HEADERS})
        for r in rows:
            name = r.get("name", "")
            writer.writerow({
                "name": name,
                "nosep": Path(name).stem,
                "status": r.get("status", ""),
                "keterangan": _STATUS_LABEL.get(r.get("status", ""), ""),
                "pages": r.get("pages", ""),
                "detail": r.get("detail", ""),
            })


def missing_report(sources: InputSources) -> tuple[list[dict], dict]:
    """Laporan nama file yang tidak lengkap antar folder sumber.

    Sumber = anak langsung folder input (mis. input/E-Klaim, input/Berkas
    Digital). Baris = nama yang tidak ada di SEMUA sumber, atau duplikat
    dalam satu sumber. Return (rows matriks, ringkasan).
    """
    files = sources.files()
    if not files:
        return [], {}
    source_names = [name for name, _ in sources.folders]
    # nama(lower) -> {sumber: jumlah}
    groups: dict[str, dict[str, int]] = {}
    orig_name: dict[str, str] = {}
    for p in files:
        src = sources.source_name(p)
        key = p.name.lower()
        orig_name.setdefault(key, p.name)
        per = groups.setdefault(key, {})
        per[src] = per.get(src, 0) + 1

    rows = []
    for name in sorted(groups):
        per = groups[name]
        ada = {s for s, c in per.items() if c > 0}
        missing = sorted(set(source_names) - ada)
        dups = sorted((s, c) for s, c in per.items() if c > 1)
        if not missing and not dups:
            continue
        row = {"nosep": Path(orig_name[name]).stem, "nama_file": orig_name[name]}
        for s in source_names:
            row[s] = "ada" if s in ada else "TIDAK"
        ket = []
        if missing:
            ket.append("tidak ada di: " + ", ".join(missing))
        if dups:
            ket.append("duplikat di: " + ", ".join(f"{s}({c}x)" for s, c in dups))
        row["keterangan"] = " | ".join(ket)
        rows.append(row)

    ringkasan = {
        "sumber": source_names,
        "per_sumber": {s: sum(1 for p in files if sources.source_name(p) == s)
                       for s in source_names},
        "nama_union": len(groups),
        "lengkap": sum(1 for per in groups.values()
                       if all(per.get(s, 0) == 1 for s in source_names)),
        "bermasalah": len(rows),
    }
    return rows, ringkasan


def _write_missing(rows: list[dict], sources: list[str], report_dir: Path) -> None:
    import csv
    out = report_dir / "combine_missing.csv"
    fieldnames = ["nosep", "nama_file"] + sources + ["keterangan"]
    labels = {"nosep": "NOSEP", "nama_file": "Nama File", "keterangan": "Keterangan"}
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writerow({k: labels.get(k, k) for k in fieldnames})
        writer.writerows(rows)


def _autosize(ws, cap: int = 40) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=4) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, cap)


def _write_xlsx(rows: list[dict], missing_rows: list[dict],
                sources: list[str], ringkasan: dict, report_dir: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        log.warning("openpyxl belum terpasang — lewati laporan Excel "
                    "(pip install openpyxl).")
        return
    fill_color = {"merged": "C6EFCE", "copied": "C6EFCE", "identical": "C6EFCE",
                  "skipped": "FFEB9C", "exists_skipped": "FFEB9C", "failed": "FFC7CE"}
    wb = Workbook()

    ws = wb.active
    ws.title = "Hasil"
    ws.append([_RESULT_HEADER_LABEL[k] for k in _RESULT_HEADERS])
    for r in rows:
        name = r.get("name", "")
        ws.append([name, Path(name).stem, r.get("status", ""),
                   _STATUS_LABEL.get(r.get("status", ""), ""),
                   r.get("pages", ""), r.get("detail", "")])
        color = fill_color.get(r.get("status"))
        if color:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=color)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)

    ws2 = wb.create_sheet("Ringkasan")
    ws2.append(["Status", "Jumlah"])
    for status in _STATUS_LABEL:
        ws2.append([_STATUS_LABEL[status],
                    sum(1 for r in rows if r.get("status") == status)])
    ws2.append(["Total", len(rows)])
    if ringkasan:
        ws2.append([])
        ws2.append(["Ringkasan Missing", ""])
        for s in ringkasan.get("sumber", []):
            ws2.append([f"File di {s}", ringkasan["per_sumber"].get(s, 0)])
        ws2.append(["Nama unik (union)", ringkasan.get("nama_union", 0)])
        ws2.append(["Lengkap (ada di semua sumber)", ringkasan.get("lengkap", 0)])
        ws2.append(["Bermasalah (missing/duplikat)", ringkasan.get("bermasalah", 0)])
    for c in ws2[1]:
        c.font = Font(bold=True)
    _autosize(ws2)

    if missing_rows:
        ws3 = wb.create_sheet("Missing")
        fieldnames = ["nosep", "nama_file"] + sources + ["keterangan"]
        labels = {"nosep": "NOSEP", "nama_file": "Nama File", "keterangan": "Keterangan"}
        ws3.append([labels.get(k, k) for k in fieldnames])
        for r in missing_rows:
            ws3.append([r.get(k, "") for k in fieldnames])
        for c in ws3[1]:
            c.font = Font(bold=True)
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = ws3.dimensions
        _autosize(ws3)

    out = report_dir / "combine_result.xlsx"
    wb.save(out)
    log.info("Laporan Excel: %s", out)


def _dedupe_per_source(paths: list[Path], sources: InputSources) -> tuple[list[Path], list[str]]:
    """Bila dalam SATU folder sumber ada beberapa salinan nama sama:
    identik -> sisakan 1; beda isi -> pilih versi terbaru (mtime).
    Return (path terpilih, catatan log)."""
    per_source: dict[str, list[Path]] = {}
    for p in paths:
        src = sources.source_name(p)
        per_source.setdefault(src, []).append(p)
    chosen: list[Path] = []
    notes: list[str] = []
    for src in sorted(per_source):
        ps = per_source[src]
        if len(ps) == 1:
            chosen.append(ps[0])
            continue
        if len({sha256(p) for p in ps}) == 1:
            chosen.append(ps[0])
            notes.append(f"duplikat identik di {src} — dipakai 1 salinan")
        else:
            newest = max(ps, key=lambda p: p.stat().st_mtime)
            chosen.append(newest)
            ts = datetime.fromtimestamp(newest.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            notes.append(f"revisi di {src}: {len(ps)} salinan beda isi — "
                         f"dipakai versi terbaru (mtime {ts})")
    return chosen, notes


def process_group(name: str, paths: list[Path], sources: InputSources, out_dir: Path,
                  only_duplicates: bool, force: bool, engine: str) -> dict:
    orig_name = paths[0].name  # pertahankan case asli
    stem = Path(orig_name).stem
    ext = Path(orig_name).suffix or ".pdf"
    out_file = out_dir / f"{stem}{ext}"

    if out_file.exists() and not force:
        return {"name": orig_name, "status": "exists_skipped", "pages": 0,
                "detail": "output sudah ada (pakai --force untuk menimpa)"}

    paths, notes = _dedupe_per_source(paths, sources)

    if len(paths) == 1:
        if only_duplicates:
            return {"name": orig_name, "status": "skipped", "pages": 0,
                    "detail": "nama unik (default: hanya duplikat)"}
        shutil.copy2(paths[0], out_file)
        return {"name": orig_name, "status": "copied", "pages": page_count(out_file),
                "detail": "; ".join(notes)}

    ordered = sorted(paths, key=lambda p: _nat_key(str(p)))
    if len({sha256(p) for p in paths}) == 1:
        detail = f"file identik dari {len(paths)} folder"
        if notes:
            detail += " | " + "; ".join(notes)
        shutil.copy2(ordered[0], out_file)
        return {"name": orig_name, "status": "identical", "pages": page_count(out_file),
                "detail": detail}

    try:
        pages = merge_any(ordered, out_file, engine)
        detail = " + ".join(sources.source_name(p) for p in ordered)
        if notes:
            detail += " | " + "; ".join(notes)
        return {"name": orig_name, "status": "merged", "pages": pages,
                "detail": detail}
    except Exception as exc:
        return {"name": orig_name, "status": "failed", "pages": 0, "detail": f"{exc}"}


# ---------------------------------------------------------------- main
def main(argv=None, progress_callback: Callable[[str, int, int], None] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Gabungkan PDF bernama sama dari subfolder berbeda (read-only pada sumber).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--input", default=str(PROJECT_DIR / "input"),
                   help="Folder sumber (dipindai rekursif)")
    p.add_argument("--eklaim-dir",
                   help="Folder E-Klaim untuk mode dua folder langsung")
    p.add_argument("--berkas-digital-dir",
                   help="Folder Berkas Digital untuk mode dua folder langsung")
    p.add_argument("--output", default=str(PROJECT_DIR / "result"),
                   help="Folder hasil")
    p.add_argument("--include-unique", action="store_true",
                   help="Proses juga nama yang hanya muncul di 1 folder "
                        "(default: hanya nama duplikat)")
    p.add_argument("--engine", choices=["auto", "pypdf", "gs"], default=COMBINE_ENGINE,
                   help="Engine gabung: auto = Ghostscript lalu fallback pypdf")
    p.add_argument("--force", action="store_true", help="Timpa output yang sudah ada")
    p.add_argument("--dry-run", action="store_true", help="Rencana saja, tanpa menulis")
    p.add_argument("--xlsx", action="store_true",
                   help="Tulis juga laporan Excel (butuh openpyxl)")
    p.add_argument("--check", action="store_true",
                   help="Cek nama file vs NOSEP di isi PDF (audit, tanpa merge)")
    p.add_argument("--safe", action="store_true",
                   help="Verifikasi nama file vs NOSEP lalu merge; file mismatch dilewati")
    p.add_argument("--force-verify", action="store_true",
                   help="Abaikan cache verifikasi (full re-check)")
    args = p.parse_args(argv)
    direct_values = (args.eklaim_dir, args.berkas_digital_dir)
    if any(direct_values) and not all(direct_values):
        p.error("--eklaim-dir dan --berkas-digital-dir harus dipakai bersamaan")
    if all(direct_values):
        eklaim_dir = Path(args.eklaim_dir)
        berkas_dir = Path(args.berkas_digital_dir)
        if not eklaim_dir.is_dir() or not berkas_dir.is_dir():
            p.error("Folder E-Klaim dan Berkas Digital harus tersedia")
        if eklaim_dir.resolve() == berkas_dir.resolve():
            p.error("Folder E-Klaim dan Berkas Digital tidak boleh sama")
        sources = InputSources((("E-Klaim", eklaim_dir), ("Berkas Digital", berkas_dir)))
    else:
        root = Path(args.input)
        if not root.is_dir():
            p.error(f"Folder input tidak ditemukan: {root}")
        sources = sources_from_root(root)
    report_dir = Path(args.output)
    global log
    log = _setup_logging(report_dir)

    blocked: set[str] = set()
    if args.check:
        if not HAS_PYPDF:
            log.error("Verifikasi nama file butuh pypdf (pip install pypdf).")
            return 2
        root = sources.folders[0][1]
        if not root.is_dir():
            log.error("Folder input tidak ditemukan: %s", root)
            return 2
        log.info("CHECK MODE — verifikasi nama file vs NOSEP di isi (tanpa merge)")
        rows, counts = verify_names(sources, args.force_verify, progress_callback)
        _write_check_csv(rows, report_dir)
        log.info("Ringkasan: total=%s ok=%s mismatch=%s no_sep=%s no_text=%s "
                 "(dari cache=%s)", counts["total"], counts["ok"],
                 counts["mismatch"], counts["no_sep"], counts["no_text"],
                 counts["cached"])
        if counts["mismatch"]:
            log.warning("ADA %s file dengan nama tidak sesuai isi! Perbaiki nama "
                        "file-nya (lihat combine_check.csv).", counts["mismatch"])
        else:
            log.info("Semua nama file cocok dengan SEP di isi. OK.")
        return 0
    if args.safe:
        if not HAS_PYPDF:
            log.error("Mode --safe butuh pypdf (pip install pypdf).")
            return 2
        root = sources.folders[0][1]
        if not root.is_dir():
            log.error("Folder input tidak ditemukan: %s", root)
            return 2
        log.info("SAFE MODE — verifikasi nama file vs NOSEP, file mismatch dilewati")
        rows, counts = verify_names(sources, args.force_verify, progress_callback)
        _write_check_csv(rows, report_dir)
        log.info("Cek: total=%s ok=%s mismatch=%s no_sep=%s no_text=%s (cache=%s)",
                 counts["total"], counts["ok"], counts["mismatch"],
                 counts["no_sep"], counts["no_text"], counts["cached"])
        blocked = {r["path"].lower()
                   for r in rows if r["status"] == "mismatch"}
        if blocked:
            log.warning("SAFE MODE: %s file mismatch DILEWATI dari merge "
                        "(lihat combine_check.csv).", len(blocked))
        else:
            log.info("Semua nama file cocok dengan isi. Lanjut merge normal.")

    log.info("=" * 60)
    log.info("Gabungkan PDF — dimulai (engine=%s)", args.engine)
    if args.engine == "pypdf" and not HAS_PYPDF:
        log.error("pypdf tidak terpasang (pip install pypdf) — tidak bisa pakai engine pypdf.")
        return 2
    if args.engine == "gs" and not find_gs():
        log.error("Ghostscript tidak ditemukan (set EKLAIM_GS_EXECUTABLE atau install gs).")
        return 2
    if args.engine == "auto" and not find_gs() and not HAS_PYPDF:
        log.error("Ghostscript dan pypdf tidak tersedia — install salah satunya.")
        return 2
    if args.engine == "auto" and find_gs():
        log.info("Ghostscript terdeteksi — dipakai sebagai engine utama.")

    out_dir = Path(args.output)

    files = sources.files()
    if blocked:
        n_before = len(files)
        files = [p for p in files if str(p.resolve()).lower() not in blocked]
        log.warning("Safe mode: %s file dilewati dari merge (%s -> %s file diproses).",
                    n_before - len(files), n_before, len(files))
        if not files:
            log.error("Semua file dilewati — tidak ada yang bisa diproses.")
            return 2
    if not files:
        log.error("Tidak ada PDF di %s", root)
        return 2
    groups = group_by_name(files)
    duplicates = {n: ps for n, ps in groups.items() if len(ps) > 1}
    log.info("Total PDF: %s | nama unik: %s | nama duplikat: %s",
             len(files), len(groups), len(duplicates))

    if args.dry_run:
        log.info("DRY RUN — tidak ada file yang ditulis. Output akan ke: %s", out_dir)
        for name in sorted(duplicates):
            paths = sorted(duplicates[name], key=lambda p: _nat_key(str(p)))
            log.info("  %s -> %s: %s", Path(name).name, len(paths),
                     " ; ".join(p.parent.name for p in paths))
        return 0

    out_dir.mkdir(parents=True, exist_ok=True)
    log.info("Output: %s", out_dir)

    rows = []
    merged = copied = failed = 0
    only_dup = not args.include_unique
    total_groups = len(groups)
    for index, name in enumerate(sorted(groups), start=1):
        if progress_callback:
            progress_callback(f"Menggabungkan {index}/{total_groups}: {Path(name).stem}", index - 1, total_groups)
        r = process_group(name, groups[name], sources, out_dir,
                          only_dup, args.force, args.engine)
        rows.append(r)
        if r["status"] == "merged":
            merged += 1
        elif r["status"] in ("copied", "identical"):
            copied += 1
        elif r["status"] == "failed":
            failed += 1
        log.info("%-16s %-14s %4s hal  %s", r["status"], r["name"][:32],
                 r["pages"] or "-", r["detail"])

    _write_result(rows, report_dir)
    missing_rows, ringkasan = missing_report(sources)
    if ringkasan:
        _write_missing(missing_rows, ringkasan["sumber"], report_dir)
        log.info("Missing report: %s nama unik (%s lengkap, %s bermasalah) — "
                 "combine_missing.csv", ringkasan["nama_union"],
                 ringkasan["lengkap"], ringkasan["bermasalah"])
    if args.xlsx:
        _write_xlsx(rows, missing_rows, ringkasan.get("sumber", []), ringkasan, report_dir)
    log.info("Selesai. merged=%s copied/identical=%s failed=%s "
             "(ringkasan: combine_result.csv)", merged, copied, failed)
    if progress_callback:
        progress_callback("Penggabungan selesai", total_groups, total_groups)
    return 0


if __name__ == "__main__":
    sys.exit(main())
